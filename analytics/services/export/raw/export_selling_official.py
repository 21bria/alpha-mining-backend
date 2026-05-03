import tempfile
from datetime import datetime

from django.core.files import File
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font

from selling.models import SellingOfficialView
from analytics.export_registry import register_exporter


def to_float(value):
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def excel_safe_datetime(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value


def _apply_filters(qs, params):
    iup_id = params.get("iup_id") or params.get("iup")
    if iup_id not in (None, "", "null", "undefined"):
        try:
            qs = qs.filter(iup_id=int(iup_id))
        except (TypeError, ValueError):
            pass

    search = params.get("search")
    if search not in (None, "", "null", "undefined"):
        qs = qs.filter(
            Q(so_number__icontains=search) |
            Q(product_code__icontains=search) |
            Q(barge_code__icontains=search) |
            Q(username__icontains=search) |
            Q(name_surveyor__icontains=search) |
            Q(factory_stock__icontains=search) |
            Q(iup_code__icontains=search) |
            Q(iup_name__icontains=search)
        )

    ordering = params.get("ordering")
    allowed_ordering = {
        "id", "-id",
        "start_date", "-start_date",
        "end_date", "-end_date",
        "created_at", "-created_at",
        "so_number", "-so_number",
        "product_code", "-product_code",
        "barge_code", "-barge_code",
        "type_selling", "-type_selling",
    }

    if ordering in allowed_ordering:
        qs = qs.order_by(ordering)
    else:
        qs = qs.order_by("-id")

    return qs


@register_exporter("selling.official")
def export_selling_official(job):
    qs = SellingOfficialView.objects.all()
    qs = _apply_filters(qs, job.params or {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Selling Official"

    headers = [
        "ID",
        "IUP ID",
        "IUP Code",
        "IUP Name",
        "Surveyor ID",
        "Surveyor Name",
        "Factory ID",
        "Factory Name",
        "Type Selling",
        "Tonnage",
        "Ni",
        "Co",
        "Al2O3",
        "CaO",
        "Cr2O3",
        "Fe",
        "MgO",
        "SiO2",
        "MnO",
        "MC",
        "SO Number",
        "Product Code",
        "Barge Code",
        "Start Date",
        "End Date",
        "Re Assay",
        "User ID",
        "Username",
        "Created At",
    ]

    ws.append(headers)

    for col in ws[1]:
        col.font = Font(bold=True)

    for row in qs:
        ws.append([
            row.id,
            row.iup_id or "",
            row.iup_code or "",
            row.iup_name or "",
            row.surveyor_id or "",
            row.name_surveyor or "",
            row.id_factory or "",
            row.factory_stock or "",
            row.type_selling or "",
            to_float(row.tonnage),
            to_float(row.ni),
            to_float(row.co),
            to_float(row.al2o3),
            to_float(row.cao),
            to_float(row.cr2o3),
            to_float(row.fe),
            to_float(row.mgo),
            to_float(row.sio2),
            to_float(row.mno),
            to_float(row.mc),
            row.so_number or "",
            row.product_code or "",
            row.barge_code or "",
            row.start_date if row.start_date else None,
            row.end_date if row.end_date else None,
            row.re_assay if row.re_assay is not None else "",
            row.user_id or "",
            row.username or "",
            excel_safe_datetime(row.created_at) if row.created_at else None,
        ])

    # format date: Start Date (24), End Date (25)
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in [24, 25]:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell.number_format = "YYYY-MM-DD"

    # format datetime: Created At (29)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=29)
        if cell.value:
            cell.number_format = "YYYY-MM-DD HH:MM:SS"

    # auto width
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    filename = f"selling_official_export_{job.id}.xlsx"
    return File(open(tmp.name, "rb"), name=filename)