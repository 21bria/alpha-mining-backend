import tempfile
from django.core.files import File
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime

from master.models import MineUnitsExportView
from analytics.export_registry import register_exporter

def excel_safe_datetime(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value

def _apply_filters(qs, params):
    iup_id = params.get("iup_id") or params.get("iup")
    if iup_id not in (None, "", "null", "undefined"):
        try:
            qs = qs.filter(iup_id=int(iup_id))
        except (TypeError, ValueError):
            pass

    search = params.get("search")
    if search not in (None, "", "null", "undefined"):
        qs = qs.filter(
            Q(unit_vendor__icontains=search) |
            Q(unit_code__icontains=search) |
            Q(unit_model__icontains=search) |
            Q(unit_class__icontains=search) |
            Q(brand__icontains=search) |
            Q(category__icontains=search) |
            Q(iup_code__icontains=search) |
            Q(iup_name__icontains=search)
        )

    ordering = params.get("ordering")
    allowed_ordering = {
        "id", "-id",
        "unit_vendor", "-unit_vendor",
        "unit_code", "-unit_code",
        "created_at", "-created_at",
    }
    if ordering in allowed_ordering:
        qs = qs.order_by(ordering)
    else:
        qs = qs.order_by("unit_vendor", "unit_code")

    return qs


@register_exporter("master.units")
def export_master_units(job):
    qs = MineUnitsExportView.objects.all()
    qs = _apply_filters(qs, job.params or {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Master Units"

    headers = [
        "ID",
        # "IUP ID",
        "IUP Code",
        "IUP Name",
        "Unit Vendor",
        "Unit Code",
        "Unit Model",
        "Unit Class",
        "Brand",
        "Category ID",
        "Category",
        "Vendor ID",
        "Supports",
        "Status",
        "Description",
        "Commisioning Date",
        "On Hire",
        "Off Hire",
        "User ID",
        "Username",
        "Assignment Start Date",
        "Assignment End Date",
        "Assignment Active",
        "Created At",
        "Updated At",
    ]
    ws.append(headers)

    for col in ws[1]:
        col.font = Font(bold=True)

    for row in qs.iterator(chunk_size=2000):
        ws.append([
            row.id,
            # row.iup_id or "",
            row.iup_code or "",
            row.iup_name or "",
            row.unit_vendor or "",
            row.unit_code or "",
            row.unit_model or "",
            row.unit_class or "",
            row.brand or "",
            row.id_category or "",
            row.category or "",
            row.id_vendor or "",
            row.supports or "",
            row.status if row.status is not None else "",
            row.description or "",
            row.commisioning_date if row.commisioning_date else None,
            row.on_hire if row.on_hire else None,
            row.off_hire if row.off_hire else None,
            row.user_id or "",
            row.username or "",
            row.assignment_start_date if row.assignment_start_date else None,
            row.assignment_end_date if row.assignment_end_date else None,
            row.assignment_active if row.assignment_active is not None else "",
            excel_safe_datetime(row.created_at) if row.created_at else None,
            excel_safe_datetime(row.updated_at) if row.updated_at else None,
        ])

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in [23, 24]:  # created_at, updated_at
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell.number_format = "YYYY-MM-DD HH:MM:SS"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    filename = f"master_units_export_{job.id}.xlsx"
    return File(open(tmp.name, "rb"), name=filename)