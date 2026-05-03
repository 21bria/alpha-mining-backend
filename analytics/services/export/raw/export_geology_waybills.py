import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import numbers
from django.core.files import File
from django.db.models import Q
from datetime import datetime, time

from geology.models import listWaybills
from analytics.export_registry import register_exporter

def excel_safe_datetime(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value


def excel_safe_time(value):
    if isinstance(value, time):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value

def _apply_filters(qs, params):
    iup_id = params.get("iup_id") or params.get("iup")
    if iup_id not in (None, "", "null", "undefined"):
        try:
            qs = qs.filter(iup_id=int(iup_id))
        except (TypeError, ValueError):
            pass

    tgl_from = params.get("date_start")
    if tgl_from not in (None, "", "null", "undefined"):
        qs = qs.filter(tgl_deliver__gte=tgl_from)

    tgl_to = params.get("date_end")
    if tgl_to not in (None, "", "null", "undefined"):
        qs = qs.filter(tgl_deliver__lte=tgl_to)

    search = params.get("search")
    if search not in (None, "", "null", "undefined"):
        qs = qs.filter(
            Q(tgl_deliver__icontains=search) |
            Q(iup_code__icontains=search) |
            Q(iup_name__icontains=search)
        )

    ordering = params.get("ordering")
    allowed_ordering = {"id", "-id", "tgl_deliver", "-tgl_deliver"}
    if ordering in allowed_ordering:
        qs = qs.order_by(ordering)
    else:
        qs = qs.order_by("tgl_deliver")

    return qs


@register_exporter("geology.waybills")
def export_geology_waybills(job):
    qs = listWaybills.objects.all()
    qs = _apply_filters(qs, job.params or {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Waybills"

    # HEADER
    ws.append([
        "IUP Code",
        "IUP Name",
        "Deliver Date",
        "Delivery Time",
        "Waybill Number",
        "Qty",
        "Sample ID",
        "Sample Status",
        "MRAL Order",
        "ROA Order",
        "Remarks",
        "Username",
    ])

    # DATA
    for row in qs.iterator(chunk_size=2000):
        ws.append([
            row.iup_code or "",
            row.iup_name or "",
            row.tgl_deliver if row.tgl_deliver else None,
            excel_safe_time(row.delivery_time) if row.delivery_time else None,
            row.waybill_number or "",
            row.qty or 0,
            row.sample_id or "",
            row.sample_status or "",
            row.mral_order or "",
            row.roa_order or "",
            row.remarks or "",
            row.username or "",
        ])

    # FORMAT DATE (kolom 3)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=3)
        if cell.value:
            cell.number_format = "YYYY-MM-DD"

    # FORMAT TIME (kolom 4)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=4)
        if cell.value:
            cell.number_format = "HH:MM:SS"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    filename = f"waybills_export_{job.id}.xlsx"
    return File(open(tmp.name, "rb"), name=filename)