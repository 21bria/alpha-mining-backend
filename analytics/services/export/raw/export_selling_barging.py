import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import numbers
from django.core.files import File
from django.db.models import Q
from datetime import datetime, time

from selling.models import SellingDetailsBargingView
from analytics.export_registry import register_exporter

def excel_safe_datetime(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value


def excel_safe_time(value):
    if isinstance(value, time):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value

def _apply_filters(qs, params):
    iup_id = params.get("iup_id") or params.get("iup")
    if iup_id not in (None, "", "null", "undefined"):
        try:
            qs = qs.filter(iup_id=int(iup_id))
        except (TypeError, ValueError):
            pass

    tgl_from = params.get("date_start")
    if tgl_from not in (None, "", "null", "undefined"):
        qs = qs.filter(date_hauling__gte=tgl_from)

    tgl_to = params.get("date_end")
    if tgl_to not in (None, "", "null", "undefined"):
        qs = qs.filter(date_hauling__lte=tgl_to)

    search = params.get("search")
    if search not in (None, "", "null", "undefined"):
        qs = qs.filter(
            Q(date__icontains=search) |
            Q(iup_code__icontains=search) |
            Q(iup_name__icontains=search)
        )

    ordering = params.get("ordering")
    allowed_ordering = {"id", "-id", "date_hauling", "-date_hauling"}
    if ordering in allowed_ordering:
        qs = qs.order_by(ordering)
    else:
        qs = qs.order_by("date_hauling")

    return qs

@register_exporter("selling.details_barging")
def export_selling_details_barging(job):
    qs = SellingDetailsBargingView.objects.all()
    qs = _apply_filters(qs, job.params or {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Selling Barging"

    # HEADER
    ws.append([
        "IUP Code",
        "IUP Name",
        "Date Barge In",
        "Date Barge Out",
        "Barge Code",
        "Shift",
        "Dome",
        "Stockpile",
        "Material",
        "Unit Code",
        "Ritase",
        "Tonnage",
        "Ton Barge Load",
        "Ton Barge Unload",
        "Fill Adjust",
        "Batch",
        "Code Increment",
        "Code Sub",
        "Code Batch In",
        "Code Batch Ex",
        "Code Batch Pulp",
        "Survey Order",
        "Code Fix Batch",
        "Code Lot",
        "Factory Stock",
        "Type Selling",
        "Date Hauling",
        "Time Hauling",
        "No Input",
        "Sale Adjust",
        "Sale Dome",
        "Status Barging",
        "Direct",
        "Description",
        "Created At",
        "Username",
    ])

    # DATA
    for row in qs.iterator(chunk_size=2000):
        ws.append([
            row.iup_code or "",
            row.iup_name or "",
            row.date_barge_in if row.date_barge_in else None,
            row.date_barge_out if row.date_barge_out else None,
            row.barge_code or "",
            row.shift or "",
            row.dome or "",
            row.stockpile or "",
            row.material or "",
            row.unit_code or "",
            row.ritase or 0,
            float(row.tonnage or 0),
            float(row.ton_barge_load or 0),
            float(row.ton_barge_unload or 0),
            float(row.fill_adjust or 0),
            row.batch or "",
            row.code_inc or "",
            row.code_sub or "",
            row.code_batch_in or "",
            row.code_batch_ex or "",
            row.code_batch_pulp or "",
            row.surv_order or "",
            row.code_fix_batch or "",
            row.code_lot or "",
            row.factory_stock or "",
            row.type_selling or "",
            row.date_hauling if row.date_hauling else None,
            excel_safe_time(row.time_hauling) if row.time_hauling else None,
            row.no_input or "",
            row.sale_adjust or "",
            row.sale_dome or "",
            row.status_barging or "",
            row.direct or "",
            row.description or "",
            excel_safe_datetime(row.created_at) if row.created_at else None,
            row.username or "",
        ])

    # FORMAT DATE (kolom 3, 4, 27)
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in [3, 4, 27]:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell.number_format = "YYYY-MM-DD"

    # FORMAT TIME (kolom 28)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=28)
        if cell.value:
            cell.number_format = "HH:MM:SS"

    # FORMAT DATETIME (kolom 35)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=35)
        if cell.value:
            cell.number_format = "YYYY-MM-DD HH:MM:SS"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    filename = f"selling_details_barging_export_{job.id}.xlsx"
    return File(open(tmp.name, "rb"), name=filename)