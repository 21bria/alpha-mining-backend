import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import numbers
from django.core.files import File
from django.db.models import Q
from datetime import datetime, time

from geology.models import AssayMral
from analytics.export_registry import register_exporter

def excel_safe_datetime(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value


def excel_safe_time(value):
    if isinstance(value, time):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value

def _apply_filters(qs, params):
    iup_id = params.get("iup_id") or params.get("iup")
    if iup_id not in (None, "", "null", "undefined"):
        try:
            qs = qs.filter(iup_id=int(iup_id))
        except (TypeError, ValueError):
            pass

    tgl_from = params.get("date_start")
    if tgl_from not in (None, "", "null", "undefined"):
        qs = qs.filter(release_date__gte=tgl_from)

    tgl_to = params.get("date_end")
    if tgl_to not in (None, "", "null", "undefined"):
        qs = qs.filter(release_date__lte=tgl_to)

    search = params.get("search")
    if search not in (None, "", "null", "undefined"):
        qs = qs.filter(
            Q(release_date__icontains=search) |
            Q(iup_code__icontains=search) |
            Q(iup_name__icontains=search)
        )

    ordering = params.get("ordering")
    allowed_ordering = {"id", "-id", "release_date", "-release_date"}
    if ordering in allowed_ordering:
        qs = qs.order_by(ordering)
    else:
        qs = qs.order_by("release_date")

    return qs

@register_exporter("lab.assay_mral")
def export_lab_assay_mral(job):
    qs = AssayMral.objects.select_related("iup", "user").all()
    qs = _apply_filters(qs, job.params or {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Assay ROA"

    # HEADER
    ws.append([
        "IUP Code",
        "IUP Name",
        "Release",
        "Job Number",
        "Sample ID",
        "Ni",
        "Co",
        "Fe2O3",
        "Fe",
        "MgO",
        "SiO2",
        "Username",
        "Created At",
    ])

    # DATA
    for row in qs.iterator(chunk_size=2000):
        ws.append([
            getattr(row.iup, "iup_code", "") or "",
            getattr(row.iup, "iup_name", "") or "",
            excel_safe_datetime(row.release_mral) if row.release_roa else None,
            row.job_number or "",
            row.sample_id or "",
            float(row.ni or 0),
            float(row.co or 0),
            float(row.fe2o3 or 0),
            float(row.fe or 0),
            float(row.mgo or 0),
            float(row.sio2 or 0),
            getattr(row.user, "username", "") or "",
            excel_safe_datetime(row.created_at) if row.created_at else None,
        ])

    # FORMAT DATE (kolom 4)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=3)
        if cell.value:
            cell.number_format = "YYYY-MM-DD HH:MM:SS"

    # FORMAT DATETIME CREATED AT (kolom 39)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=13)
        if cell.value:
            cell.number_format = "YYYY-MM-DD HH:MM:SS"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    filename = f"assay_mral_export_{job.id}.xlsx"
    return File(open(tmp.name, "rb"), name=filename)