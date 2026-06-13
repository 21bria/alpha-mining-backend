import tempfile
from datetime import datetime

from django.core.files import File
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font

from geology.models import SamplesDomeView
from analytics.export_registry import register_exporter


def to_float(value):
    try:
        if value in (None, "", "null", "undefined"):
            return None
        return float(value)
    except Exception:
        return None


def excel_safe_datetime(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value


def _apply_filters(qs, params):
    iup_id = params.get("iup_id") or params.get("iup")
    if iup_id not in (None, "", "null", "undefined"):
        try:
            qs = qs.filter(iup_id=int(iup_id))
        except (TypeError, ValueError):
            pass

    sampling_point = params.get("sampling_point")
    date_from = params.get("date_from")
    date_to = params.get("date_to")

    # Sama seperti filter page:
    # kalau dome dipilih, tanggal diabaikan
    if sampling_point not in (None, "", "null", "undefined", "all"):
        qs = qs.filter(sampling_point__iexact=sampling_point)
    else:
        if date_from not in (None, "", "null", "undefined"):
            qs = qs.filter(date_sample__gte=date_from)

        if date_to not in (None, "", "null", "undefined"):
            qs = qs.filter(date_sample__lte=date_to)

    search = params.get("search")
    if search not in (None, "", "null", "undefined"):
        qs = qs.filter(
            Q(sample_id__icontains=search) |
            Q(material__icontains=search) |
            Q(batch__icontains=search) |
            Q(type_sample__icontains=search) |
            Q(sample_method__icontains=search) |
            Q(sampling_point__icontains=search) |
            Q(sampling_desc__icontains=search) |
            Q(username__icontains=search) |
            Q(iup_code__icontains=search) |
            Q(iup_name__icontains=search)
        )

    ordering = params.get("ordering")
    allowed_ordering = {
        "id", "-id",
        "date_sample", "-date_sample",
        "sample_id", "-sample_id",
        "material", "-material",
        "sampling_point", "-sampling_point",
        "batch", "-batch",
        "created_at", "-created_at",
    }

    if ordering in allowed_ordering:
        qs = qs.order_by(ordering)
    else:
        qs = qs.order_by("-date_sample", "sample_id")

    return qs


@register_exporter("geology.sample_dome")
def export_geology_sample_dome(job):
    qs = SamplesDomeView.objects.all()
    qs = _apply_filters(qs, job.params or {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Sample Dome"

    headers = [
        # "ID",
        # "IUP ID",
        "IUP Code",
        "IUP Name",
        "Date Sample",
        "Shift",
        "Type Sample",
        "Category",
        "Sample Method",
        "Material",
        "Sampling Point",
        "Batch",
        "Increments",
        "Sample Weight",
        "Sample ID",
        "Sampling Desc",
        "Ni",
        "Co",
        "Al2O3",
        "CaO",
        "Cr2O3",
        "Fe2O3",
        "Fe",
        "MgO",
        "SiO2",
        "MC",
        "SM",
        # "User ID",
        # "Username",
        "Created At",
    ]

    ws.append(headers)

    for col in ws[1]:
        col.font = Font(bold=True)

    for row in qs:
        ws.append([
            # str(row.id) if row.id else "",
            # row.iup_id or "",
            row.iup_code or "",
            row.iup_name or "",
            row.date_sample if row.date_sample else None,
            row.shift or "",
            row.type_sample or "",
            row.category or "",
            row.sample_method or "",
            row.material or "",
            row.sampling_point or "",
            row.batch or "",
            row.increments if row.increments is not None else "",
            to_float(row.sample_weight),
            row.sample_id or "",
            row.sampling_desc or "",
            to_float(row.ni),
            to_float(row.co),
            to_float(row.al2o3),
            to_float(row.cao),
            to_float(row.cr2o3),
            to_float(row.fe2o3),
            to_float(row.fe),
            to_float(row.mgo),
            to_float(row.sio2),
            to_float(row.mc),
            to_float(row.sm),
            # row.user_id or "",
            # row.username or "",
            excel_safe_datetime(row.created_at) if row.created_at else None,
        ])

    # Date Sample column = 5
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=5)
        if cell.value:
            cell.number_format = "YYYY-MM-DD"

    # FORMAT DATETIME (Created At kolom 26)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=26)
        if cell.value:
            cell.number_format = 'YYYY-MM-DD HH:MM:SS'
                 
    # auto width
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            try:
                max_length = max(max_length, len(str(cell.value or "")))
            except Exception:
                pass

        ws.column_dimensions[column_letter].width = min(max_length + 2, 13)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    filename = f"sample_dome_export_{job.id}.xlsx"
    return File(open(tmp.name, "rb"), name=filename)