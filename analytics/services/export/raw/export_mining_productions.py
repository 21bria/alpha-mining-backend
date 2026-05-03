import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import numbers
from django.core.files import File
from django.db.models import Q

from mining.models import mineProductionsView
from analytics.export_registry import register_exporter

def _apply_filters(qs, params):
    iup_id = params.get("iup_id") or params.get("iup")
    if iup_id not in (None, "", "null", "undefined"):
        try:
            qs = qs.filter(iup_id=int(iup_id))
        except (TypeError, ValueError):
            pass

    tgl_from = params.get("date_start")
    if tgl_from not in (None, "", "null", "undefined"):
        qs = qs.filter(date_production__gte=tgl_from)

    tgl_to = params.get("date_end")
    if tgl_to not in (None, "", "null", "undefined"):
        qs = qs.filter(date_production__lte=tgl_to)

    search = params.get("search")
    if search not in (None, "", "null", "undefined"):
        qs = qs.filter(
            Q(date_production__icontains=search) |
            Q(iup_code__icontains=search) |
            Q(iup_name__icontains=search)
        )

    ordering = params.get("ordering")
    allowed_ordering = {"id", "-id", "date_production", "-date_production"}
    if ordering in allowed_ordering:
        qs = qs.order_by(ordering)
    else:
        qs = qs.order_by("date_production")

    return qs

@register_exporter("mining.productions")
def export_mining_productions(job):
    qs = mineProductionsView.objects.all()
    qs = _apply_filters(qs, job.params or {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Mining Productions"

    # HEADER
    ws.append([
        "IUP Code",
        "IUP Name",
        "Date Production",
        "Shift",
        "Vendors",
        "Loader",
        "Loader Class",
        "Bucket",
        "Hauler",
        "Source Area",
        "Loading Point",
        "Dumping Point",
        "Dome ID",
        "Category Mine",
        "Time Loading",
        "Time Dumping",
        "Mine Block",
        "From RL",
        "To RL",
        "RL",
        "Material",
        "Ritase",
        "BCM",
        "Tonnage",
        "Direct",
        "Remarks",
        "Created At",
        "Username",
    ])

    # DATA
    for row in qs.iterator(chunk_size=2000):
        ws.append([
            row.iup_code or "",
            row.iup_name or "",
            row.date_production if row.date_production else None,
            row.shift or "",
            row.vendors or "",
            row.loader or "",
            row.hauler_class or "",
            row.bucket or 0,
            row.hauler or "",
            # row.hauler_type or "",
            row.sources_area or "",
            row.loading_point or "",
            row.dumping_point or "",
            row.dome_id or "",
            row.category_mine or "",
            str(row.time_loading or ""),
            str(row.time_dumping or ""),
            row.mine_block or "",
            row.from_rl or "",
            row.to_rl or "",
            row.rl or "",
            row.nama_material or "",
            row.ritase or 0,
            float(row.bcm or 0),
            float(row.tonnage or 0),
            row.direct or "",
            row.remarks or "",
            str(row.created_at or ""),
            row.username or "",
        ])

    # FORMAT DATE (kolom ke-3)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=3)
        if cell.value:
            cell.number_format = 'YYYY-MM-DD'

    # FORMAT TIME LOADING (kolom ke-15)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=15)
        if cell.value:
            cell.number_format = 'HH:MM:SS'

    # FORMAT TIME DUMPING (kolom ke-16)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=16)
        if cell.value:
            cell.number_format = 'HH:MM:SS'

    # FORMAT DATETIME (Created At kolom 27)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=27)
        if cell.value:
            cell.number_format = 'YYYY-MM-DD HH:MM:SS'
                 
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    filename = f"mining_production_export_{job.id}.xlsx"
    return File(open(tmp.name, "rb"), name=filename)