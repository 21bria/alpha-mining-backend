import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import numbers
from django.core.files import File
from django.db.models import Q
from datetime import datetime, time

from geology.models import SamplesView
from analytics.export_registry import register_exporter

def excel_safe_datetime(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value


def excel_safe_time(value):
    if isinstance(value, time):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value


def _apply_filters(qs, params):
    # exclude selling
    qs = qs.exclude(category__iexact="selling")

    iup_id = params.get("iup_id") or params.get("iup")
    if iup_id not in (None, "", "null", "undefined"):
        try:
            qs = qs.filter(iup_id=int(iup_id))
        except (TypeError, ValueError):
            pass

    tgl_from = params.get("date_start")
    if tgl_from not in (None, "", "null", "undefined"):
        qs = qs.filter(date_sample__gte=tgl_from)

    tgl_to = params.get("date_end")
    if tgl_to not in (None, "", "null", "undefined"):
        qs = qs.filter(date_sample__lte=tgl_to)

    search = params.get("search")
    if search not in (None, "", "null", "undefined"):
        qs = qs.filter(
            Q(date_sample__icontains=search) |
            Q(iup_code__icontains=search) |
            Q(iup_name__icontains=search) |
            Q(sample_id__icontains=search)
        )

    ordering = params.get("ordering")
    allowed_ordering = {"id", "-id", "date_sample", "-date_sample"}
    if ordering in allowed_ordering:
        qs = qs.order_by(ordering)
    else:
        qs = qs.order_by("date_sample")

    return qs

@register_exporter("geology.sample")
def export_geology_sample(job):
    qs = SamplesView.objects.all()
    qs = _apply_filters(qs, job.params or {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Geology Samples"

    # HEADER
    ws.append([
        "IUP Code",
        "IUP Name",
        "Date Sample",
        "Shift",
        "Type Sample",
        "Category",
        "Sample Method",
        "Material",
        "Sampling Area",
        "Sampling Point",
        "Batch",
        "Increments",
        "Size",
        "Sample Weight",
        "Sample ID",
        "Remark",
        "Primer Raw",
        "Duplicate Raw",
        "Sampling Desc",
        "Created At",
        "Username",
    ])

    # DATA
    for row in qs.iterator(chunk_size=2000):
        ws.append([
            row.iup_code or "",
            row.iup_name or "",
            row.date_sample if row.date_sample else None,
            row.shift or "",
            row.type_sample or "",
            row.category or "",
            row.sample_method or "",
            row.material or "",
            row.sampling_area or "",
            row.sampling_point or "",
            row.batch or "",
            row.increments or 0,
            row.size or "",
            float(row.sample_weight or 0),
            row.sample_id or "",
            row.remark or "",
            float(row.primer_raw or 0),
            float(row.duplicate_raw or 0),
            row.sampling_desc or "",
            excel_safe_datetime(row.created_at) if row.created_at else None,
            row.username or "",
        ])

    # FORMAT DATE (kolom 3)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=3)
        if cell.value:
            cell.number_format = "YYYY-MM-DD"

    # FORMAT DATETIME (kolom 26)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=20)
        if cell.value:
            cell.number_format = "YYYY-MM-DD HH:MM:SS"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    filename = f"geology_sample_export_{job.id}.xlsx"
    return File(open(tmp.name, "rb"), name=filename)