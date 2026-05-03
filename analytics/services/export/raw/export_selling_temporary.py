import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import numbers
from django.core.files import File
from django.db.models import Q
from datetime import datetime, time

from selling.models import SellingBargingTemporaryView
from analytics.export_registry import register_exporter

def excel_safe_datetime(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value


def excel_safe_time(value):
    if isinstance(value, time):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value


def _apply_filters(qs, params):
    iup_id = params.get("iup_id") or params.get("iup")
    if iup_id not in (None, "", "null", "undefined"):
        try:
            qs = qs.filter(iup_id=int(iup_id))
        except (TypeError, ValueError):
            pass

    tgl_from = params.get("date_start")
    if tgl_from not in (None, "", "null", "undefined"):
        qs = qs.filter(date_hauling__gte=tgl_from)

    tgl_to = params.get("date_end")
    if tgl_to not in (None, "", "null", "undefined"):
        qs = qs.filter(date_hauling__lte=tgl_to)

    search = params.get("search")
    if search not in (None, "", "null", "undefined"):
        qs = qs.filter(
            Q(date__icontains=search) |
            Q(iup_code__icontains=search) |
            Q(iup_name__icontains=search)
        )

    ordering = params.get("ordering")
    allowed_ordering = {"id", "-id", "date_hauling", "-date_hauling"}
    if ordering in allowed_ordering:
        qs = qs.order_by(ordering)
    else:
        qs = qs.order_by("date_hauling")

    return qs


@register_exporter("selling.barging_temporary")
def export_selling_barging_temporary(job):
    qs = SellingBargingTemporaryView.objects.all()
    qs = _apply_filters(qs, job.params or {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Selling Barging Temporary"

    # HEADER
    ws.append([
        "IUP Code",
        "IUP Name",
        "Date Hauling",
        "Time Hauling",
        "Barge Code",
        "Shift",
        "Dome",
        "Stockpile",
        "Material",
        "Unit Code",
        "Tonnage",
        "Code Lot",
        "Code Increment",
        "Code Sub",
        "Type Selling",
        "Sale Adjust",
        "No Urut",
        "Description",
        "Status",
        "Created At",
        "Username",
    ])

    # DATA
    for row in qs.iterator(chunk_size=2000):
        ws.append([
            row.iup_code or "",
            row.iup_name or "",
            row.date_hauling if row.date_hauling else None,
            excel_safe_time(row.time_hauling) if row.time_hauling else None,
            row.barge_code or "",
            row.shift or "",
            row.dome or "",
            row.stockpile or "",
            row.material or "",
            row.unit_code or "",
            float(row.tonnage or 0),
            row.code_lot or "",
            row.code_inc or "",
            row.code_sub or "",
            row.type_selling or "",
            row.sale_adjust or "",
            row.no_urut or 0,
            row.description or "",
            row.status or 0,
            excel_safe_datetime(row.created_at) if row.created_at else None,
            row.username or "",
        ])

    # FORMAT DATE (kolom 3)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=3)
        if cell.value:
            cell.number_format = "YYYY-MM-DD"

    # FORMAT TIME (kolom 4)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=4)
        if cell.value:
            cell.number_format = "HH:MM:SS"

    # FORMAT DATETIME (kolom 20)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=20)
        if cell.value:
            cell.number_format = "YYYY-MM-DD HH:MM:SS"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    filename = f"selling_barging_temporary_export_{job.id}.xlsx"
    return File(open(tmp.name, "rb"), name=filename)