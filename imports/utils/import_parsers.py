import csv
import re
from pathlib import Path
from typing import List, Dict, Any

import openpyxl


def normalize_header(value: str) -> str:
    value = str(value or "").strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_")


def apply_alias(headers: list[str], aliases: dict | None):
    if not aliases:
        return headers
    return [aliases.get(h, h) for h in headers]


def validate_headers(headers, required_headers=None, allowed_headers=None):
    headers = [h for h in headers if h]

    header_set = set(headers)

    if required_headers:
        missing = set(required_headers) - header_set
        if missing:
            raise ValueError(
                f"Missing required columns: {', '.join(sorted(missing))}"
            )

    if allowed_headers:
        unknown = header_set - set(allowed_headers)
        if unknown:
            raise ValueError(
                f"Unknown columns: {', '.join(sorted(unknown))}"
            )


def read_csv(
    path: str,
    required_headers=None,
    allowed_headers=None,
    aliases=None,
) -> List[Dict[str, Any]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)

        raw_headers = reader.fieldnames or []
        headers = [normalize_header(h) for h in raw_headers]
        headers = apply_alias(headers, aliases)

        validate_headers(headers, required_headers, allowed_headers)

        rows = []
        for r in reader:
            item = {}
            for k, v in r.items():
                key = normalize_header(k)
                key = aliases.get(key, key) if aliases else key
                item[key] = v
            if any(v is not None and str(v).strip() != "" for v in item.values()):
                rows.append(item)

        return rows


def read_xlsx(
    path: str,
    required_headers=None,
    allowed_headers=None,
    aliases=None,
) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    raw_headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_header(h) for h in raw_headers]
    headers = apply_alias(headers, aliases)

    validate_headers(headers, required_headers, allowed_headers)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            item[h] = row[i] if i < len(row) else None

        if any(v is not None and str(v).strip() != "" for v in item.values()):
            rows.append(item)

    return rows


def read_file(
    path: str,
    required_headers=None,
    allowed_headers=None,
    aliases=None,
) -> List[Dict[str, Any]]:
    ext = Path(path).suffix.lower()

    if ext == ".csv":
        return read_csv(
            path,
            required_headers=required_headers,
            allowed_headers=allowed_headers,
            aliases=aliases,
        )

    if ext in [".xlsx", ".xlsm"]:
        return read_xlsx(
            path,
            required_headers=required_headers,
            allowed_headers=allowed_headers,
            aliases=aliases,
        )

    raise ValueError(f"Unsupported file type: {ext}")