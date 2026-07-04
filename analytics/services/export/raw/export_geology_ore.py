import os
import tempfile
from openpyxl import Workbook
from django.core.files import File
from django.db.models import Q

from geology.models import OreProductionsView
from analytics.export_registry import register_exporter

def _apply_filters(qs, params):
    iup_id = params.get("iup_id") or params.get("iup")
    if iup_id not in (None, "", "null", "undefined"):
        try:
            qs = qs.filter(iup_id=int(iup_id))
        except (TypeError, ValueError):
            pass

    tgl_from = params.get("date_start")
    if tgl_from not in (None, "", "null", "undefined"):
        qs = qs.filter(tgl_production__gte=tgl_from)

    tgl_to = params.get("date_end")
    if tgl_to not in (None, "", "null", "undefined"):
        qs = qs.filter(tgl_production__lte=tgl_to)

    search = params.get("search")
    if search not in (None, "", "null", "undefined"):
        qs = qs.filter(
            Q(tgl_production__icontains=search) |
            Q(iup_code__icontains=search) |
            Q(iup_name__icontains=search)
        )

    ordering = params.get("ordering")
    allowed_ordering = {"id", "-id", "tgl_production", "-tgl_production"}
    if ordering in allowed_ordering:
        qs = qs.order_by(ordering)
    else:
        qs = qs.order_by("tgl_production")

    return qs


@register_exporter("geology.ore")
def export_geology_ore(job):
    qs = OreProductionsView.objects.all()
    qs = _apply_filters(qs, job.params or {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Ore Productions"

    ws.append([
        # "ID",
        "IUP Code",
        "IUP Name",
        "Tanggal Production",
        "Category",
        "Shift",
        "Prospect Area",
        "Pit Dome",
        "Mine Block",
        "From RL",
        "To RL",
        "Material",
        "Ore Class",
        "Ni Grade",
        "Grade Control",
        "Unit Truck",
        "Stockpile",
        "Pile ID",
        "Batch Code",
        "Increment",
        "Batch Status",
        "Ritase",
        "Tonnage",
        "Pile Status",
        "Truck Factor",
        "Remarks",
        "Sample Number",
        "Direct",
        "Created At",
        "Username",
    ])

    for row in qs.iterator(chunk_size=2000):
        ws.append([
            # str(row.id) if row.id else "",
            row.iup_code or "",
            row.iup_name or "",
            # str(row.tgl_production or ""),
            row.tgl_production if row.tgl_production else None,
            row.category or "",
            row.shift or "",
            row.prospect_area or "",
            row.pit_dome or "",
            row.mine_block or "",
            row.from_rl or "",
            row.to_rl or "",
            row.nama_material or "",
            row.ore_class or "",
            float(row.ni_grade or 0),
            row.grade_control or "",
            row.unit_truck or "",
            row.stockpile or "",
            row.pile_id or "",
            row.batch_code or "",
            row.increment or 0,
            row.batch_status or "",
            row.ritase or 0,
            float(row.tonnage or 0),
            row.pile_status or "",
            row.truck_factor or "",
            row.remarks or "",
            row.sample_number or "",
            row.direct or "",
            str(row.created_at or ""),
            row.username or "",
        ])
    # FORMAT DATE (kolom ke-3)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=3)
        if cell.value:
            cell.number_format = 'YYYY-MM-DD'

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    filename = f"geology_ore_export_{job.id}.xlsx"
    return File(open(tmp.name, "rb"), name=filename)