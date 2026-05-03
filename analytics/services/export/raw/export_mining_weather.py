import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import numbers
from django.core.files import File
from django.db.models import Q
from datetime import datetime, time

from mining.models import Weather
from analytics.export_registry import register_exporter

def excel_safe_datetime(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value


def excel_safe_time(value):
    if isinstance(value, time):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value

def _apply_filters(qs, params):
    iup_id = params.get("iup_id") or params.get("iup")
    if iup_id not in (None, "", "null", "undefined"):
        try:
            qs = qs.filter(iup_id=int(iup_id))
        except (TypeError, ValueError):
            pass

    tgl_from = params.get("date_start")
    if tgl_from not in (None, "", "null", "undefined"):
        qs = qs.filter(date__gte=tgl_from)

    tgl_to = params.get("date_end")
    if tgl_to not in (None, "", "null", "undefined"):
        qs = qs.filter(date__lte=tgl_to)

    search = params.get("search")
    if search not in (None, "", "null", "undefined"):
        qs = qs.filter(
            Q(date__icontains=search) |
            Q(iup_code__icontains=search) |
            Q(iup_name__icontains=search)
        )

    ordering = params.get("ordering")
    allowed_ordering = {"id", "-id", "date", "-date"}
    if ordering in allowed_ordering:
        qs = qs.order_by(ordering)
    else:
        qs = qs.order_by("date")

    return qs



@register_exporter("mining.weather")
def export_mining_weather(job):
    qs = Weather.objects.select_related("iup", "user").all()
    qs = _apply_filters(qs, job.params or {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Weather"

    # HEADER
    ws.append([
        "IUP Code",
        "IUP Name",
        "Date",
        "Shift",
        "Category",
        "Start Time",
        "End Time",
        "Duration",
        "Description",
        "Created At",
        "Username",
    ])

    # DATA
    for row in qs.iterator(chunk_size=2000):
        ws.append([
            getattr(row.iup, "iup_code", "") or "",
            getattr(row.iup, "iup_name", "") or "",
            row.date if row.date else None,
            row.shift or "",
            row.category or "",
            excel_safe_time(row.start_time) if row.start_time else None,
            excel_safe_time(row.end_time) if row.end_time else None,
            float(row.duration or 0),
            row.description or "",
            excel_safe_datetime(row.created_at) if getattr(row, "created_at", None) else None,
            getattr(row.user, "username", "") or "",
        ])

    # FORMAT DATE (kolom 3)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=3)
        if cell.value:
            cell.number_format = "YYYY-MM-DD"

    # FORMAT TIME START (kolom 6)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=6)
        if cell.value:
            cell.number_format = "HH:MM:SS"

    # FORMAT TIME END (kolom 7)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=7)
        if cell.value:
            cell.number_format = "HH:MM:SS"

    # FORMAT DATETIME CREATED AT (kolom 10)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=10)
        if cell.value:
            cell.number_format = "YYYY-MM-DD HH:MM:SS"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    filename = f"weather_export_{job.id}.xlsx"
    return File(open(tmp.name, "rb"), name=filename)