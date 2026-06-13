import tempfile
from datetime import datetime, time

from django.core.files import File
from django.db.models import Q

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from mining.models import PlanProduction
from analytics.export_registry import register_exporter


MATERIAL_HEADERS = [
    "Top Soil",
    "OB",
    # "LGLO",
    # "MGLO",
    # "HGLO",
    "Waste",
    "Spoil",
    # "LGSO",
    # "UGLO",
    # "MGSO",
    # "HGSO",
    "Quarry",
    "Ballast",
    "Biomass",
    "MWS",
    "LIM",
    "SAP",
]


def excel_safe_datetime(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value


def excel_safe_time(value):
    if isinstance(value, time):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value


def normalize_material_name(value):
    if value is None:
        return ""

    return (
        str(value)
        .strip()
        .upper()
        .replace("_", " ")
        .replace("-", " ")
    )


def build_details_map(details):
    result = {}

    for detail in details:
        key = normalize_material_name(
            detail.material_name or detail.material_code
        )

        if not key:
            continue

        result[key] = result.get(key, 0) + float(detail.tonnage or 0)

    return result


def get_material_value(details_map, material_name):
    key = normalize_material_name(material_name)
    return float(details_map.get(key, 0))


def _apply_filters(qs, params):
    iup_id = params.get("iup_id") or params.get("iup")

    if iup_id not in (None, "", "null", "undefined"):
        try:
            qs = qs.filter(iup_id=int(iup_id))
        except (TypeError, ValueError):
            pass

    tgl_from = params.get("date_start")

    if tgl_from not in (None, "", "null", "undefined"):
        qs = qs.filter(date_plan__gte=tgl_from)

    tgl_to = params.get("date_end")

    if tgl_to not in (None, "", "null", "undefined"):
        qs = qs.filter(date_plan__lte=tgl_to)

    search = params.get("search")

    if search not in (None, "", "null", "undefined"):
        qs = qs.filter(
            Q(date_plan__icontains=search)
            | Q(category__icontains=search)
            | Q(source_code__icontains=search)
            | Q(vendor_code__icontains=search)
            | Q(iup__iup_code__icontains=search)
            | Q(iup__iup_name__icontains=search)
            | Q(details__material_name__icontains=search)
            | Q(details__material_code__icontains=search)
        ).distinct()

    ordering = params.get("ordering")

    allowed_ordering = {
        "id",
        "-id",
        "date_plan",
        "-date_plan",
        "category",
        "-category",
        "source_code",
        "-source_code",
        "vendor_code",
        "-vendor_code",
    }

    if ordering in allowed_ordering:
        qs = qs.order_by(ordering)
    else:
        qs = qs.order_by("date_plan")

    return qs


@register_exporter("mining.plan_productions")
def export_mining_plan_productions(job):
    qs = (
        PlanProduction.objects
        .select_related("iup", "user")
        .prefetch_related("details")
        .all()
    )

    qs = _apply_filters(qs, job.params or {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Mining Plan Productions"

    headers = [
        "IUP Code",
        "IUP Name",
        "Date Plan",
        "Category",
        "Sources",
        "Vendors",
        *MATERIAL_HEADERS,
        "Total",
        # "Ref Plan",
        # "Task ID",
        "Created At",
        "Username",
    ]

    ws.append(headers)

    # STYLE HEADER
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # DATA
    for row in qs.iterator(chunk_size=1000):
        details = list(row.details.all())
        details_map = build_details_map(details)

        material_values = [
            get_material_value(details_map, material_name)
            for material_name in MATERIAL_HEADERS
        ]

        total = sum(material_values)

        ws.append([
            getattr(row.iup, "iup_code", "") or "",
            getattr(row.iup, "iup_name", "") or "",
            row.date_plan if row.date_plan else None,
            row.category or "",
            row.source_code or "",
            row.vendor_code or "",
            *material_values,
            total,
            # row.ref_plan or "",
            # row.task_id or "",
            excel_safe_datetime(row.created_at)
            if getattr(row, "created_at", None)
            else None,
            getattr(row.user, "username", "") or "",
        ])

    # FORMAT DATE
    date_col = headers.index("Date Plan") + 1
    created_at_col = headers.index("Created At") + 1

    for row_idx in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row_idx, column=date_col)

        if date_cell.value:
            date_cell.number_format = "YYYY-MM-DD"

        created_cell = ws.cell(row=row_idx, column=created_at_col)

        if created_cell.value:
            created_cell.number_format = "YYYY-MM-DD HH:MM:SS"

    # FORMAT NUMBER COLUMNS
    first_material_col = headers.index(MATERIAL_HEADERS[0]) + 1
    total_col = headers.index("Total") + 1

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(first_material_col, total_col + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"

    # FREEZE + FILTER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # AUTO WIDTH
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            max_length = max(max_length, len(str(value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 28)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    filename = f"mining_plan_productions_export_{job.id}.xlsx"

    return File(open(tmp.name, "rb"), name=filename)