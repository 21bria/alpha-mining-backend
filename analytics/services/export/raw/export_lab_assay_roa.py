import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import numbers
from django.core.files import File
from django.db.models import Q
from datetime import datetime, time

from geology.models import AssayRoa
from analytics.export_registry import register_exporter

def excel_safe_datetime(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value


def excel_safe_time(value):
    if isinstance(value, time):
        return value.replace(tzinfo=None) if value.tzinfo else value
    return value

def _apply_filters(qs, params):
    iup_id = params.get("iup_id") or params.get("iup")
    if iup_id not in (None, "", "null", "undefined"):
        try:
            qs = qs.filter(iup_id=int(iup_id))
        except (TypeError, ValueError):
            pass

    tgl_from = params.get("date_start")
    if tgl_from not in (None, "", "null", "undefined"):
        qs = qs.filter(release_date__gte=tgl_from)

    tgl_to = params.get("date_end")
    if tgl_to not in (None, "", "null", "undefined"):
        qs = qs.filter(release_date__lte=tgl_to)

    search = params.get("search")
    if search not in (None, "", "null", "undefined"):
        qs = qs.filter(
            Q(release_date__icontains=search) |
            Q(iup_code__icontains=search) |
            Q(iup_name__icontains=search)
        )

    ordering = params.get("ordering")
    allowed_ordering = {"id", "-id", "release_date", "-release_date"}
    if ordering in allowed_ordering:
        qs = qs.order_by(ordering)
    else:
        qs = qs.order_by("release_date")

    return qs

@register_exporter("lab.assay_roa")
def export_lab_assay_roa(job):
    qs = AssayRoa.objects.select_related("iup", "user").all()
    qs = _apply_filters(qs, job.params or {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Assay ROA"

    # HEADER
    ws.append([
        # "IUP ID",
        "IUP Code",
        "IUP Name",
        # "Release Date",
        # "Release Time",
        "Release",
        "Job Number",
        "Sample ID",
        "Ni",
        "Co",
        "Al2O3",
        "CaO",
        "Cr2O3",
        "Fe2O3",
        "Fe",
        # "K2O",
        "MgO",
        "MnO",
        # "Na2O",
        # "P2O5",
        # "P",
        "SiO2",
        # "TiO2",
        # "S",
        # "Cu",
        # "Zn",
        # "Cl",
        # "SO3",
        # "LOI",
        # "Total",
        # "WT Wet",
        # "WT Dry",
        "MC",
        # "P75um",
        # ">5mm",
        # "Problem",
        # "No Input",
        "Username",
        "Created At",
    ])

    # DATA
    for row in qs.iterator(chunk_size=2000):
        ws.append([
            # row.iup_id or "",
            getattr(row.iup, "iup_code", "") or "",
            getattr(row.iup, "iup_name", "") or "",
            # row.release_date if row.release_date else None,
            # excel_safe_time(row.release_time) if row.release_time else None,
            excel_safe_datetime(row.release_roa) if row.release_roa else None,
            row.job_number or "",
            row.sample_id or "",
            float(row.ni or 0),
            float(row.co or 0),
            float(row.al2o3 or 0),
            float(row.cao or 0),
            float(row.cr2o3 or 0),
            float(row.fe2o3 or 0),
            float(row.fe or 0),
            # float(row.k2o or 0),
            float(row.mgo or 0),
            float(row.mno or 0),
            # float(row.na2o or 0),
            # float(row.p2o5 or 0),
            # float(row.p or 0),
            float(row.sio2 or 0),
            # float(row.tio2 or 0),
            # float(row.s or 0),
            # float(row.cu or 0),
            # float(row.zn or 0),
            # float(row.ci or 0),
            # float(row.so3 or 0),
            # float(row.loi or 0),
            # row.total or "",
            # float(row.wt_wet or 0),
            # float(row.wt_dry or 0),
            float(row.mc or 0),
            # float(row.p75um or 0),
            # float(row._5mm or 0),
            # row.problem or "",
            # row.no_input or 0,
            getattr(row.user, "username", "") or "",
            excel_safe_datetime(row.created_at) if row.created_at else None,
        ])

    # FORMAT DATE (kolom 4)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=3)
        if cell.value:
            cell.number_format = "YYYY-MM-DD HH:MM:SS"

    # # FORMAT TIME (kolom 5)
    # for row_idx in range(2, ws.max_row + 1):
    #     cell = ws.cell(row=row_idx, column=5)
    #     if cell.value:
    #         cell.number_format = "HH:MM:SS"

    # # FORMAT DATETIME (kolom 6)
    # for row_idx in range(2, ws.max_row + 1):
    #     cell = ws.cell(row=row_idx, column=6)
    #     if cell.value:
    #         cell.number_format = "YYYY-MM-DD HH:MM:SS"

    # FORMAT DATETIME CREATED AT (kolom 39)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=18)
        if cell.value:
            cell.number_format = "YYYY-MM-DD HH:MM:SS"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    filename = f"assay_roa_export_{job.id}.xlsx"
    return File(open(tmp.name, "rb"), name=filename)